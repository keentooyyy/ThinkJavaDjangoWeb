import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl.utils import get_column_letter

from GameProgress.services.ranking import get_all_student_rankings
from StudentManagementSystem.decorators.custom_decorators import session_login_required
from StudentManagementSystem.models import Student
from StudentManagementSystem.models.roles import Role
from StudentManagementSystem.models.teachers import HandledSection
from StudentManagementSystem.views.ranking_view import get_common_params


@session_login_required(role=[Role.ADMIN, Role.TEACHER])  # ✅ Only Admin & Teacher allowed
def export_ranking_xls(request):
    params = get_common_params(request)
    user = request.user_obj

    # Default: no student restriction
    limit_to_students = None

    # Restrict to handled students if Teacher
    if user.role == Role.TEACHER:
        handled_sections = HandledSection.objects.filter(teacher=user)
        student_ids = Student.objects.filter(
            section__in=handled_sections.values_list("section_id", flat=True)
        ).values_list("id", flat=True)

        limit_to_students = student_ids

    # Fetch rankings
    rankings = get_all_student_rankings(
        sort_by=params["sort_by"],
        sort_order=params["sort_order"],
        filter_by=params["section_filter"],
        department_filter=None if not params["department_name"] or params["department_name"].lower() == "all"
                                else params["department_name"],
        limit_to_students=limit_to_students,
    )

    # ✅ Apply search filter
    search_query = request.GET.get("search", "").strip().lower()
    if search_query:
        rankings = [
            r for r in rankings
            if search_query in str(r.get("student_id", "")).lower()
               or search_query in str(r.get("first_name", "")).lower()
               or search_query in str(r.get("last_name", "")).lower()
               or search_query in f"{r.get('first_name', '')} {r.get('last_name', '')}".lower()
               or search_query in str(r.get("section", "")).lower()
               or search_query in str(r.get("score", "")).lower()
        ]

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Rankings"

    # Headers
    headers = [
        "Student ID",
        "First Name",
        "Last Name",
        "Department & Section",
        "Time Remaining",
        "Achievements",
        "Score",
    ]
    ws.append(headers)

    # Rows
    for r in rankings:
        ws.append([
            r.get("student_id", ""),
            r.get("first_name", ""),
            r.get("last_name", ""),
            r.get("section", ""),
            r.get("total_time_remaining", ""),
            r.get("achievements_unlocked", ""),
            r.get("score", ""),
        ])

    # Auto column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="student_rankings.xlsx"'

    wb.save(response)
    return response


@session_login_required(role=[Role.ADMIN, Role.TEACHER])
def print_ranking(request):
    params = get_common_params(request)
    user = request.user_obj

    # Default: no student restriction
    limit_to_students = None

    # If teacher, restrict to their handled sections only
    if user.role == Role.TEACHER:
        from StudentManagementSystem.models.teachers import HandledSection
        from StudentManagementSystem.models import Student

        handled_sections = HandledSection.objects.filter(teacher=user)
        student_ids = Student.objects.filter(
            section__in=handled_sections.values_list("section_id", flat=True)
        ).values_list("id", flat=True)

        limit_to_students = student_ids

    # Fetch rankings
    rankings = get_all_student_rankings(
        sort_by=params["sort_by"],
        sort_order=params["sort_order"],
        filter_by=params["section_filter"],
        department_filter=None if not params["department_name"] or params["department_name"].lower() == "all"
                                else params["department_name"],
        limit_to_students=limit_to_students,
    )

    # ✅ Apply search filter if provided
    search_query = request.GET.get("search", "").strip().lower()
    if search_query:
        rankings = [
            r for r in rankings
            if search_query in str(r.get("student_id", "")).lower()
               or search_query in str(r.get("first_name", "")).lower()
               or search_query in str(r.get("last_name", "")).lower()
               or search_query in f"{r.get('first_name', '')} {r.get('last_name', '')}".lower()
               or search_query in str(r.get("section", "")).lower()
               or search_query in str(r.get("score", "")).lower()
        ]

    context = {
        "rankings": rankings,
        "selected_department": params["department_name"],
        "selected_section": params["section_filter"],
    }

    return render(request, "print_layout.html", context)

